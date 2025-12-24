from openpyxl import load_workbook

PATH = "/home/ds/Downloads/Code_frames.xlsx"

wb = load_workbook(PATH)
print("Sheets:", wb.sheetnames)

ws = wb[wb.sheetnames[0]]   # TEMP: first sheet explicitly

COL_MANUAL = 4    # D
COL_AUTO   = 8    # H
COL_RESULT = 18   # R

changed = 0

for r in range(2, ws.max_row + 1):
    manual_val = ws.cell(r, COL_MANUAL).value
    auto_val   = ws.cell(r, COL_AUTO).value

    if manual_val is None or auto_val is None:
        continue

    try:
        manual = int(str(manual_val).strip())
    except:
        continue

    auto_str = str(auto_val).strip()

    if "-" in auto_str:
        try:
            start, end = map(int, auto_str.split("-"))
        except:
            continue
    else:
        try:
            start = end = int(auto_str)
        except:
            continue

    if start <= manual <= end:
        result = 0
    else:
        result = min(abs(manual - start), abs(manual - end))

    ws.cell(r, COL_RESULT).value = result
    changed += 1

wb.save(PATH)
print("Updated rows:", changed)
print("DONE.")